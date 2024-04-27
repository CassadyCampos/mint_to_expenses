import os
import pandas as pd
import boto3
import json

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from io import StringIO

def perform_tangerine_transformation(input_filename, filecontents):
    input_dir = "transactions"
    output_dir = "transformed"
    try:
        input_filepath = os.path.join(input_dir, input_filename)
        output_filepath = '/tmp/' + input_filename.replace(".csv", "_transformed.xlsx")

        wb = Workbook()
        ws = wb.active

        headers = ["", "Item", "Date", "Amount (CAD)", "Split", "Category"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        rowIndex = 2
        for index, row in df.iterrows():
            item = row["Name"]
            date = pd.to_datetime(row["Transaction Date"]).date()
            amount_cad = float(row["Amount"])
            decided_split = f"=D{rowIndex}/2"
            rowIndex = rowIndex + 1
            ws.append(["", item, date, amount_cad, decided_split, category])

        for column_cells in ws.column:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = min((max_length + 2) * 1.2, 35)
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        for _ in range(3):
            ws.append([])

        ws.append(["Total", None, None, "=SUM(D2:D" + str(rowIndex) + ")", "=SUM(E2:E" + str(rowIndex) + ")"])
        ws.append([])
        wb.save(output_filepath)

        print(f"Completed processing {input_filename}. Saved {output_filepath} to storage")
        return output_filepath
    except Exception as e:
        print(f"An error occured during transformation: {str(e)}")
        return None


def perform_transformations(input_filename, fileContents):
    # Input and output directories
    input_dir = "transactions"
    output_dir = "transformed"
    try:
        input_filepath = os.path.join(input_dir, input_filename)
        output_filepath = '/tmp/' + input_filename.replace(".csv", "_transformed.xlsx")

        # Read the CSV file
        stringIO = StringIO(fileContents)
        df = pd.read_csv(stringIO)

        # Create a new Excel workbook and select the active sheet
        wb = Workbook()
        ws = wb.active

        # Write the transformed headers to the Excel sheet
        headers = ["", "Item", "Date", "Amount (CAD)", "Decided Split", "Category"]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Perform transformations and write to Excel sheet
        rowIndex = 2
        for index, row in df.iterrows():
            if row["Category"] in [
                "Transfer",
                "Deposit",
                "Credit Card Payment",
                "Hide from Budgets & Trends",
                "Bank Fee",
                "Income",
                "Investments",
                "Interest Income",
                "Mortgage & Rent",
                "Parking",
                "TFSA Investment",
                "Subscriptions"
                "Mobile Phone",
                "Books",
                "Video Games",
                "Canada Student Loan",
                "Alberta Student Loan"]:
                continue

            item = row["Description"]
            date = pd.to_datetime(row["Date"]).date()
            amount_cad = float(row["Amount"])
            decided_split = f"=D{rowIndex}/2"  # Referencing Amount (CAD) cell
            category = row["Category"]
            rowIndex = rowIndex + 1
            ws.append(["",item, date, amount_cad, decided_split, category])

        # Auto-fit column widths
        for column_cells in ws.columns:
            max_length = max(len(str(cell.value)) for cell in column_cells)
            adjusted_width = min((max_length + 2) * 1.2, 35)  # Adding some padding and scaling
            column_letter = get_column_letter(column_cells[0].column)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add three rows of space at the bottom
        for _ in range(3):
            ws.append([])

        # Add "Total" row
        ws.append(["Total", None, None, "=SUM(D2:D" + str(rowIndex) + ")", "=SUM(E2:E" + str(rowIndex) + ")"])
        # Add blank row
        ws.append([])
        # Save the Excel workbook
        wb.save(output_filepath)

        print(f"Completed processing {input_filename}. Saved {output_filepath} to storage")

        return output_filepath
    except Exception as e:
        print(f"An error occurred during transformations: {str(e)}")
        return None

def handler (event, context):
    try: 
        is_dev_env = os.environ.get('IS_DEV', False)
        if (is_dev_env):
            print("Environment is dev")
            # Access AWS credentials from environment variables
            aws_access_key_id = os.environ.get('AWS_ACCESS_KEY_ID')
            aws_secret_access_key = os.environ.get('AWS_SECRET_ACCESS_KEY')
            aws_region = os.environ.get('AWS_DEFAULT_REGION')

            s3 = boto3.client(
                's3',
                aws_access_key_id=aws_access_key_id,
                aws_secret_access_key=aws_secret_access_key,
                region_name='ca-central-1'
            )
        else:
            print("Environment is not dev")
            s3 = boto3.client(
            's3',
            )

        s3 = boto3.client('s3')
        # Define your input and output bucket names
        input_bucket_name = 'mint-transactions'
        output_bucket_name = 'mint-transformed'
        # List all objects in the input S3 bucket
        prefix = 'transactions/'

        response = s3.list_objects_v2(Bucket=input_bucket_name, Prefix=prefix)
        # print(f"Contents: {response.get('Contents', [])}" )

        print(f"Going through object item in bucket: {input_bucket_name}")

        for obj in response.get('Contents', []):
            object_key = obj['Key']
            file_name = os.path.basename(object_key)
            
            if not file_name.lower().endswith('.csv'):
                print(f"Skipping this object in bucket: {file_name}")
                continue

            # Download CSV file from input bucket
            # Extract the file name using os.path.basename
            if file_name.lower().endswith('.csv'):
               response = s3.get_object(Bucket=input_bucket_name, Key=object_key)
               print(f"Processing file: {file_name}") 
               csv_content = response['Body'].read().decode('utf-8')

               if 'tangerine' in file_name.lower():
                    print(f"Processing tangerine file: {file_name}")
                    transformed_data = perform_tangerine_transformation(file_name, csv_content)
            else:
                    transformed_data = perform_transformations(file_name, csv_content)
                
            newFileName = file_name.replace("_transactions.csv", "_transformed.xlsx")
            s3.upload_file(transformed_data, input_bucket_name, f'transformed/{newFileName}')
            print(f"Saved to S3")

        response = {
            "statusCode": 200,
            "body": json.dumps({"message": "Done processing all objects successfully!"})
        }

        try:
            print("**********************************************************************")
            print("Attempting to email subscribers...")
            sns_client = boto3.client('sns')
            topic_arn = os.environ.get('SNS_TOPIC')

            # Publish to SNS
            sns_client.publish(
                TopicArn=topic_arn,
                Message="Done processing all objects successfully!\n Download the transactions on S3 and apply any changes you'd like manually.",
                Subject='mint_to_expenses Execution Notification',
            )
            print("Subscribers emailed successfully!")
        except Exception as error:
            print(f"Unable to publish to SNS topic: {error}")

        return response
        # Rest of your code...
    except Exception as e:
        print(f"An error occurred during AWS setup: {str(e)}")