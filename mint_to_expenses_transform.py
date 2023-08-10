import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter

# Input and output directories
input_dir = "transactions"
output_dir = "transformed"

# Create the output directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

# List CSV files in the input directory
input_files = [file for file in os.listdir(input_dir) if file.endswith(".csv")]

# Process each input file
for input_filename in input_files:
    input_filepath = os.path.join(input_dir, input_filename)
    output_filepath = os.path.join(output_dir, input_filename.replace(".csv", "_transformed.xlsx"))

    # Read the CSV file
    df = pd.read_csv(input_filepath)

    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active

    # # Define a named style for currency formatting
    # currency_style = NamedStyle(name='currency', number_format='$#,##0.00')

    # # Apply the currency style to columns E, F, G, and H
    # for col_letter in ['E', 'F', 'G', 'H']:
    #     for cell in ws[col_letter]:
    #         cell.style = currency_style

    # Write the transformed headers to the Excel sheet
    headers = ["", "Item", "Date", "Paid By", "Amount (CAD)", "CherryOwesCass", "CassOwesCherry", "Split 50/50", "Category"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)


    # Perform transformations and write to Excel sheet
    rowIndex = 2
    for index, row in df.iterrows():
        if row["Category"] in [
            "Transfer",
            "Deposit",
            "Credit Card Payment",
            "Hide from Budgets & Trends",
            "Bank Fee",
            "Income",
            "Interest Income",
            "Mortgage & Rent",
            "Parking",
            "TFSA Investment",
            "Mobile Phone",
            "Books",
            "Video Games",
            "Canada Student Loan",
            "Alberta Student Loan"]:
            continue


        item = row["Description"]
        date = pd.to_datetime(row["Date"]).date()
        paid_by = "Cassady"
        amount_cad = float(row["Amount"])
        cherry_owes_cass = f"=E{rowIndex}/2"  # Referencing Amount (CAD) cell
        cass_owes_cherry = ""
        split_50_50 = f"=E{rowIndex}/2"  # Referencing Amount (CAD) cell
        category = row["Category"]

        rowIndex = rowIndex + 1
        ws.append(["",item, date, paid_by, amount_cad, cherry_owes_cass, cass_owes_cherry, split_50_50, category])

    # Auto-fit column widths
    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = min((max_length + 2) * 1.2, 50)  # Adding some padding and scaling
        column_letter = get_column_letter(column_cells[0].column)

        ws.column_dimensions[column_letter].width = adjusted_width

    # Add three rows of space at the bottom
    for _ in range(3):
        # rowIndex = rowIndex + 1
        ws.append([])

    # Add "Total" row
    ws.append(["Total", None, None, None, "=SUM(E2:E" + str(rowIndex) + ")", "=SUM(F2:F" + str(rowIndex) + ")", None, "=SUM(H2:H" + str(rowIndex)  +")"])

    # Add blank row
    ws.append([])

    # Add "TotalOwed" column
    ws.append(["", "", "", "", "TotalOwed", "=E" + str(rowIndex + 3) + " - F" + str(rowIndex + 3)])

    # Save the Excel workbook
    wb.save(output_filepath)

    print(f"Data from {input_filename} transformed and saved to {output_filepath}")
